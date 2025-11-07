import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Protection
from PIL import Image
import pytesseract
import re
from datetime import datetime
import os
import io
import shutil

# Template path
TEMPLATE_PATH = "Gasmet Reference Limits -Chemical Burning.xlsx"

# Expected components in exact order
EXPECTED_COMPONENTS = [
    "Water Vapour", "Carbon Dioxide", "Carbon Monoxide", "Nitrous Oxide",
    "Acrolein", "Phenol", "Styrene", "M-Xylene", "P-Xylene", "O-Xylene",
    "Ammonia", "Benzene", "Crotonaldehyde", "Formaldehyde", "Hydrogen Chloride",
    "Hydrogen Fluoride", "Naphthalene", "Ethyl Benzene", "Toluene", "Ethylene"
]

# Initialize session state
if "extracted_data" not in st.session_state:
    st.session_state.extracted_data = pd.DataFrame(columns=["Component", "Concentration"])
if "upload_count" not in st.session_state:
    st.session_state.upload_count = 0
if "working_file" not in st.session_state:
    st.session_state.working_file = None
if "last_uploaded_file" not in st.session_state:
    st.session_state.last_uploaded_file = None

st.set_page_config(
    page_title="Excel OCR Populator",
    page_icon="https://www.sgcleaningxpert.com/wp-content/uploads/2017/10/NEA-Logo1.png",
    layout="wide"
)

st.title("SG NEA Gasmet OCR Populator")
st.markdown("---")


# ==================== FUNCTIONS ====================

def clear_columns(excel_path, columns=['B', 'C']):
    """Clear specified columns in the Excel template."""
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        
        cleared_count = 0
        for col in columns:
            # Clear rows 2-21 (20 chemical components)
            for row_idx in range(2, 22):
                cell = ws[f"{col}{row_idx}"]
                if cell.value is not None:
                    cell.value = None
                    cleared_count += 1
        
        wb.save(excel_path)
        wb.close()
        return True, cleared_count
    except Exception as e:
        return False, str(e)


def create_working_copy():
    """Create a timestamped copy of the template with exact formatting."""
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        directory = os.path.dirname(TEMPLATE_PATH)
        filename = os.path.basename(TEMPLATE_PATH)
        name, ext = os.path.splitext(filename)
        
        new_filename = f"{name}_{timestamp}{ext}"
        new_path = os.path.join(directory, new_filename)
        
        # Use shutil.copy2 to preserve all metadata and formatting
        shutil.copy2(TEMPLATE_PATH, new_path)
        
        return new_path
    except Exception as e:
        st.error(f"Error creating working copy: {e}")
        return None


def extract_text_from_image(image):
    """Extract text from uploaded image using optimized OCR."""
    try:
        # Convert to PIL Image if needed
        if not isinstance(image, Image.Image):
            image = Image.open(image)
        
        # Optimize image for OCR
        # Convert to grayscale
        image = image.convert('L')
        
        # Enhance contrast
        from PIL import ImageEnhance
        enhancer = ImageEnhance.Contrast(image)
        image = enhancer.enhance(2.0)
        
        # Perform OCR with custom config for better accuracy
        custom_config = r'--oem 3 --psm 6'
        text = pytesseract.image_to_string(image, config=custom_config)
        
        return text
    except Exception as e:
        st.error(f"OCR Error: {e}")
        return ""


def normalize_component_name(name):
    """Normalize component name for matching."""
    if not name:
        return ""
    return str(name).strip().lower().replace(" ", "").replace("-", "").replace("_", "")


def parse_extracted_text(text):
    """Parse OCR text to extract all 20 component-concentration pairs."""
    lines = text.strip().split('\n')
    
    # Create normalized map of expected components
    expected_map = {normalize_component_name(comp): comp for comp in EXPECTED_COMPONENTS}
    
    # Dictionary to store results - use dict to avoid duplicates
    results_dict = {}
    
    # Multiple patterns to catch different formats
    patterns = [
        r'([A-Za-z\s\(\)\-]+?)[\s\.:]+([0-9,.]+)',  # Component ... 123
        r'([A-Za-z\s\(\)\-]+?)\s+([0-9,.]+)\s*(?:ppm)?',  # Component 123 ppm
        r'([A-Za-z\s\(\)\-]+?)[:\s]+([0-9,.]+)',  # Component: 123
        r'([A-Za-z\s\(\)\-]{3,})\s*([0-9]+\.?[0-9]*)',  # More flexible pattern
    ]
    
    for line in lines:
        line = line.strip()
        if not line or len(line) < 3:
            continue
        
        for pattern in patterns:
            matches = re.finditer(pattern, line)
            for match in matches:
                component = match.group(1).strip()
                concentration = match.group(2).strip()
                
                # Skip if component is too short or contains common OCR noise
                if len(component) < 3:
                    continue
                if component.lower() in ['the', 'and', 'for', 'ppm', 'reading', 'st', 'nd', 'rd', 'th', 'min']:
                    continue
                
                # Normalize and try to match
                normalized = normalize_component_name(component)
                
                # Direct match
                if normalized in expected_map:
                    matched_component = expected_map[normalized]
                    if matched_component not in results_dict:
                        results_dict[matched_component] = concentration
                    continue
                
                # Fuzzy match - check if it's a substring or close match
                for exp_norm, exp_orig in expected_map.items():
                    if exp_orig in results_dict:  # Already matched
                        continue
                    
                    # Check various matching strategies
                    match_found = False
                    
                    # Strategy 1: Check if extracted contains expected or vice versa
                    if normalized in exp_norm or exp_norm in normalized:
                        if len(normalized) >= len(exp_norm) * 0.5:  # At least 50% match
                            results_dict[exp_orig] = concentration
                            match_found = True
                            break
                    
                    # Strategy 2: Check word-by-word match for multi-word components
                    component_words = set(component.lower().split())
                    expected_words = set(exp_orig.lower().split())
                    if component_words and expected_words:
                        overlap = len(component_words & expected_words)
                        if overlap > 0 and overlap >= len(expected_words) * 0.5:
                            results_dict[exp_orig] = concentration
                            match_found = True
                            break
                    
                    if match_found:
                        break
    
    # Convert dict to list of dicts for DataFrame
    data = [{"Component": comp, "Concentration": conc} for comp, conc in results_dict.items()]
    
    # If we didn't get all 20, add placeholders for missing ones
    matched_components = set(results_dict.keys())
    missing_components = set(EXPECTED_COMPONENTS) - matched_components
    
    if missing_components:
        st.warning(f"⚠️ {len(missing_components)} component(s) not auto-detected. Please add manually:")
        for comp in sorted(missing_components):
            st.caption(f"   • {comp}")
    
    return pd.DataFrame(data)


def check_column_status(excel_path):
    """Check which columns (B or C) have data."""
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        
        b_has_data = False
        c_has_data = False
        
        # Check rows 2-21 (20 chemicals)
        for row_idx in range(2, 22):
            if ws[f"B{row_idx}"].value is not None:
                b_has_data = True
            if ws[f"C{row_idx}"].value is not None:
                c_has_data = True
            
            if b_has_data and c_has_data:
                break
        
        wb.close()
        return b_has_data, c_has_data
    except Exception as e:
        st.error(f"Error checking columns: {e}")
        return False, False


def populate_column(excel_path, data_df, target_column):
    """Populate specified column with extracted data while preserving ALL formatting including conditional formatting."""
    try:
        # Load workbook without data_only to preserve formulas and conditional formatting
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Create a mapping of normalized component names to concentrations
        data_map = {}
        for _, row in data_df.iterrows():
            if pd.notna(row["Component"]) and str(row["Component"]).strip():
                key = normalize_component_name(row["Component"])
                value = str(row["Concentration"]).strip()
                data_map[key] = value
        
        updates = 0
        # Match components in column A with our data
        # Starting from row 2 (row 1 is header)
        for row_idx in range(2, 22):  # Rows 2-21 (20 chemicals)
            component_cell = ws[f"A{row_idx}"]
            if component_cell.value:
                # Normalize the component name from Excel
                component_key = normalize_component_name(component_cell.value)
                
                if component_key in data_map:
                    # Get target cell
                    target_cell = ws[f"{target_column}{row_idx}"]
                    
                    # Convert concentration to number if possible (for conditional formatting to work)
                    value = data_map[component_key]
                    try:
                        # Remove any text like "ppm" and commas, then convert to float
                        clean_value = re.sub(r'[^\d.]', '', value)
                        if clean_value:
                            numeric_value = float(clean_value)
                            target_cell.value = numeric_value
                        else:
                            target_cell.value = value
                    except (ValueError, AttributeError):
                        # If not a number, store as string
                        target_cell.value = value
                    
                    updates += 1
        
        # Save the workbook (openpyxl automatically preserves conditional formatting)
        wb.save(excel_path)
        wb.close()
        return True, updates
    except Exception as e:
        return False, str(e)


# ==================== UI ====================

# Section 1: Clear Template
st.header("1️⃣ Clear Template Columns")
col1, col2 = st.columns([3, 1])

with col1:
    st.info(f"📁 Template: `{os.path.basename(TEMPLATE_PATH)}`")

with col2:
    if st.button("🗑️ Clear Columns B & C", type="secondary"):
        if os.path.exists(TEMPLATE_PATH):
            success, result = clear_columns(TEMPLATE_PATH)
            if success:
                st.success(f"✅ Cleared {result} cells from columns B & C")
                st.session_state.upload_count = 0  # Reset upload count
                st.session_state.extracted_data = pd.DataFrame(columns=["Component", "Concentration"])
            else:
                st.error(f"❌ Error: {result}")
        else:
            st.error("❌ Template file not found!")

st.markdown("---")

# ==================== 2️⃣ DRAG & DROP (WORKS EVERYWHERE – FINAL SOLUTION) ====================
st.header("2️⃣ Drag & Drop Your Screenshot Here")

st.markdown("""
<p style="font-size:18px; color:#d4380d;">
    <strong>Windows:</strong> Win+Shift+S → select table → drag the image here<br>
    <strong>Mac:</strong> ⌘+Shift+5 → select → drag the floating thumbnail here
</p>
""", unsafe_allow_html=True)

# THIS IS THE ONLY METHOD THAT WORKS RELIABLY
drag_drop = st.file_uploader(
    "DRAG YOUR SCREENSHOT DIRECTLY INTO THIS BOX",
    type=["png", "jpg", "jpeg", "bmp", "tiff"],
    label_visibility="collapsed",
    help="After taking screenshot, drag the image file or floating preview here"
)

if drag_drop:
    image = Image.open(drag_drop)
    st.image(image, use_column_width=True)
    
    with st.spinner("Extracting data..."):
        text = extract_text_from_image(image)
        if text.strip():
            df = parse_extracted_text(text)
            if not df.empty:
                st.session_state.extracted_data = df
                st.success(f"Extracted {len(df)} components!")
                st.balloons()
            else:
                st.warning("No data found – try clearer image")
        else:
            st.error("OCR failed")
            
# Section 3: Edit Extracted Data
st.header("3️⃣ Review & Edit Extracted Data")

if not st.session_state.extracted_data.empty:
    st.info("✏️ Review and edit the extracted data if needed")
    
    # Show expected vs extracted
    col1, col2 = st.columns(2)
    with col1:
        st.metric("Expected Components", len(EXPECTED_COMPONENTS))
    with col2:
        st.metric("Extracted Components", len(st.session_state.extracted_data))
    
    # Editable data editor
    edited_df = st.data_editor(
        st.session_state.extracted_data,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "Component": st.column_config.SelectboxColumn(
                "Component Name",
                options=EXPECTED_COMPONENTS,
                width="medium",
                required=True
            ),
            "Concentration": st.column_config.TextColumn("Concentration Value", width="medium")
        },
        hide_index=True
    )
    
    # Update session state with edited data
    st.session_state.extracted_data = edited_df
else:
    st.warning("📝 No data extracted yet. Upload an image above.")
    
    # Allow manual entry
    if st.button("➕ Add Manual Entry"):
        new_row = pd.DataFrame([{"Component": "", "Concentration": ""}])
        st.session_state.extracted_data = pd.concat(
            [st.session_state.extracted_data, new_row], 
            ignore_index=True
        )
        st.rerun()

st.markdown("---")

# Section 4: Populate Excel
# ==================== 4️⃣ POPULATE EXCEL (UPDATED) ====================
st.header("4️⃣ Populate Excel File")

if not st.session_state.extracted_data.empty:
    # --------------------------------------------------------------
    # 1. Always create a *new* timestamped copy BEFORE populating
    # --------------------------------------------------------------
    with st.spinner("Creating new timestamped file..."):
        # Create a fresh copy of the ORIGINAL template
        new_file = create_working_copy()
        if not new_file:
            st.error("Failed to create working copy.")
            st.stop()

        # If there was a previous working file, copy its filled data (B or C) into the new file
        if st.session_state.working_file and os.path.exists(st.session_state.working_file):
            try:
                # Load old and new workbooks
                old_wb = load_workbook(st.session_state.working_file)
                new_wb = load_workbook(new_file)
                old_ws = old_wb.active
                new_ws = new_wb.active

                # Copy values from columns B and C (rows 2–21) if they exist
                for row_idx in range(2, 22):
                    for col in ["B", "C"]:
                        old_cell = old_ws[f"{col}{row_idx}"]
                        new_cell = new_ws[f"{col}{row_idx}"]
                        if old_cell.value is not None:
                            # Preserve numeric type for conditional formatting
                            try:
                                new_cell.value = float(old_cell.value)
                            except (ValueError, TypeError):
                                new_cell.value = old_cell.value

                new_wb.save(new_file)
                new_wb.close()
                old_wb.close()
                st.info(f"Copied previous data from `{os.path.basename(st.session_state.working_file)}`")
            except Exception as e:
                st.warning(f"Could not copy previous data: {e}")

        # Update session state
        st.session_state.working_file = new_file

    # --------------------------------------------------------------
    # 2. Determine target column in the *new* file
    # --------------------------------------------------------------
    b_has, c_has = check_column_status(new_file)

    if not b_has:
        target_column = "B"
        reading_txt = "1st reading"
    elif not c_has:
        target_column = "C"
        reading_txt = "2nd reading"
    else:
        st.warning("Both columns already have data in the new file.")
        target_column = st.radio("Select target column:", ["B", "C"], horizontal=True)
        reading_txt = "1st reading" if target_column == "B" else "2nd reading"

    # --------------------------------------------------------------
    # 3. UI
    # --------------------------------------------------------------
    col1, col2, col3 = st.columns([2, 2, 1])
    with col1:
        st.metric("Components Ready", len(st.session_state.extracted_data))
    with col2:
        st.metric("Target Column", f"{target_column} ({reading_txt})")
    with col3:
        populate_btn = st.button("Populate Excel", type="primary")

    # --------------------------------------------------------------
    # 4. Populate
    # --------------------------------------------------------------
    if populate_btn:
        with st.spinner("Populating data..."):
            success, updates = populate_column(
                new_file,
                st.session_state.extracted_data,
                target_column
            )
            if success:
                st.session_state.upload_count += 1
                st.success(f"Populated {updates} rows → **Column {target_column}**")
                st.success(f"New file: `{os.path.basename(new_file)}`")

                # Download button
                with open(new_file, "rb") as f:
                    st.download_button(
                        label="Download Populated Excel File",
                        data=f.read(),
                        file_name=os.path.basename(new_file),
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary"
                    )

                if b_has and c_has:
                    st.success("Both readings complete!")
                else:
                    st.info("Upload another image to fill the remaining column.")
            else:
                st.error(f"Population failed: {updates}")

else:
    st.warning("No data available to populate. Upload an image first.")
    
# Footer
st.markdown("---")
st.caption("💡 **Note:** Make sure Tesseract OCR is installed on your system")
st.caption("   • Mac: `brew install tesseract`")
st.caption("   • Linux: `sudo apt-get install tesseract-ocr`")
st.caption("   • Windows: Download from [GitHub](https://github.com/UB-Mannheim/tesseract/wiki)")
