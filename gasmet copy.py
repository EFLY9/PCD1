import io
import re
import cv2
import numpy as np
import pytesseract
import pandas as pd
import openpyxl
import streamlit as st
from PIL import Image

# -----------------------------------------------------------
#  Streamlit Page Config
# -----------------------------------------------------------
st.set_page_config(page_title="Gas Analysis → Excel Populator", layout="wide")
st.title("Gas Analysis → Excel Populator (Manual Editable Version)")

st.markdown("""
### ⚙️ Steps
1. Upload your **Gasmet Excel template (.xlsx)**.  
2. Click **🧹 Clear Columns B & C** to reset values.  
3. Upload a **Gasmet result image** → OCR runs automatically.  
4. Review or manually edit the extracted table.  
5. Click **🧾 Populate Excel** to write values into column B (if blank) or C.
""")

# -----------------------------------------------------------
#  Excel Helpers
# -----------------------------------------------------------
def load_workbook_bytes(uploaded_file_bytes: bytes):
    """Load workbook from uploaded bytes."""
    return openpyxl.load_workbook(io.BytesIO(uploaded_file_bytes), data_only=False)

def clear_columns_B_C(ws):
    """Clear all non-formula values in columns B and C."""
    max_row = ws.max_row
    for r in range(1, max_row + 1):
        for col in ("B", "C"):
            cell = ws[f"{col}{r}"]
            if isinstance(cell.value, str) and cell.value.strip().startswith("="):
                continue
            cell.value = None

# -----------------------------------------------------------
#  OCR (robust local Tesseract)
# -----------------------------------------------------------
def extract_table_from_image(image_bytes: bytes):
    """Detect table, deskew, OCR, and parse Component–Concentration pairs (auto invert + fallback)."""
    image_array = np.frombuffer(image_bytes, np.uint8)
    img = cv2.imdecode(image_array, cv2.IMREAD_COLOR)

    # ---------- Step 1: Preprocess ----------
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    gray = cv2.GaussianBlur(gray, (3, 3), 0)

    # Contrast enhancement
    gray = cv2.convertScaleAbs(gray, alpha=1.5, beta=10)

    # ---------- Step 2: Edge detection ----------
    edges = cv2.Canny(gray, 50, 150, apertureSize=3)

    # ---------- Step 3: Find largest rectangular contour ----------
    contours, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    table_contour = None
    max_area = 0
    for cnt in contours:
        area = cv2.contourArea(cnt)
        if area > max_area:
            peri = cv2.arcLength(cnt, True)
            approx = cv2.approxPolyDP(cnt, 0.02 * peri, True)
            if len(approx) == 4:
                table_contour = approx
                max_area = area

    # ---------- Step 4: Deskew (if contour found) ----------
    if table_contour is not None:
        pts = table_contour.reshape(4, 2)
        rect = np.zeros((4, 2), dtype="float32")
        s = pts.sum(axis=1)
        rect[0] = pts[np.argmin(s)]
        rect[2] = pts[np.argmax(s)]
        diff = np.diff(pts, axis=1)
        rect[1] = pts[np.argmin(diff)]
        rect[3] = pts[np.argmax(diff)]

        (tl, tr, br, bl) = rect
        widthA = np.linalg.norm(br - bl)
        widthB = np.linalg.norm(tr - tl)
        heightA = np.linalg.norm(tr - br)
        heightB = np.linalg.norm(tl - bl)
        maxWidth = int(max(widthA, widthB))
        maxHeight = int(max(heightA, heightB))

        dst = np.array([
            [0, 0],
            [maxWidth - 1, 0],
            [maxWidth - 1, maxHeight - 1],
            [0, maxHeight - 1]], dtype="float32")

        M = cv2.getPerspectiveTransform(rect, dst)
        warped = cv2.warpPerspective(img, M, (maxWidth, maxHeight))
    else:
        warped = img.copy()

    # ---------- Step 5: Threshold and check for inversion ----------
    gray = cv2.cvtColor(warped, cv2.COLOR_BGR2GRAY)
    gray = cv2.fastNlMeansDenoising(gray, h=25)

    # Adaptive threshold
    binary = cv2.adaptiveThreshold(
        gray, 255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY,
        31, 10
    )

    # Detect if background is dark (invert if needed)
    if np.mean(binary) < 127:
        binary = cv2.bitwise_not(binary)

    # Sharpen edges slightly
    kernel = np.array([[0, -1, 0],
                       [-1, 5, -1],
                       [0, -1, 0]])
    binary = cv2.filter2D(binary, -1, kernel)

    # ---------- Step 6: OCR ----------
    text = pytesseract.image_to_string(binary, config="--psm 6")

    # Fallback if empty text — try the original image
    if not text.strip():
        text = pytesseract.image_to_string(img, config="--psm 6")

    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]

    # ---------- Step 7: Parse results ----------
    table = {}
    for line in lines:
        # normalize separators
        line = line.replace("=", " ").replace(":", " ").replace(",", ".")
        m = re.search(r'([A-Za-z0-9\-\s/()]+?)\s+([0-9]+(?:\.[0-9]+)?)$', line)
        if m:
            comp = m.group(1).strip()
            conc = m.group(2).strip()
            table[comp] = conc
        else:
            nums = re.findall(r'[0-9]+(?:\.[0-9]+)?', line)
            if nums:
                conc = nums[-1]
                comp = line.replace(conc, "").strip()
                if len(comp) > 1:
                    table[comp] = conc

    return table, text
# -----------------------------------------------------------
#  Streamlit UI
# -----------------------------------------------------------
uploaded_excel = st.file_uploader("Upload Excel template", type=["xlsx"])

if uploaded_excel:
    raw_bytes = uploaded_excel.read()
    wb = load_workbook_bytes(raw_bytes)
    ws = wb.active
    st.success(f"✅ Loaded sheet: {ws.title}")

    # --- Clear Columns B & C ---
    if st.button("🧹 Clear Columns B & C"):
        clear_columns_B_C(ws)
        st.success("Columns B and C cleared (formulas kept).")

    # --- Image upload & OCR ---
    uploaded_img = st.file_uploader("📸 Upload Gasmet result image", type=["png","jpg","jpeg"])
    if uploaded_img:
        image_bytes = uploaded_img.read()
        st.info("Running local OCR...")
        table, raw_text = extract_table_from_image(image_bytes)
        st.text_area("📄 Raw OCR text (for review)", raw_text or "(no text detected)", height=300)

        # --- Editable table ---
        df = pd.DataFrame(list(table.items()), columns=["Component","Concentration"])
        st.markdown("### ✏️ Review / Edit extracted data")
        edited_df = st.data_editor(df, num_rows="dynamic", use_container_width=True, key="editor_df")
        st.session_state["edited_df"] = edited_df

        # --- Populate Excel ---
        if st.button("🧾 Populate Excel with Edited Values", key="populate_btn"):
            ed = st.session_state.get("edited_df")
            if ed is None or ed.empty:
                st.error("⚠️ No rows to populate — please edit or add data above.")
            else:
                edited_map = {
                    str(r["Component"]).strip().lower().replace(" ",""): str(r["Concentration"])
                    for _, r in ed.iterrows() if str(r["Component"]).strip() != ""
                }
                updates = 0
                for r in range(1, ws.max_row + 1):
                    comp = ws[f"A{r}"].value
                    if not comp:
                        continue
                    key = str(comp).strip().lower().replace(" ","")
                    if key not in edited_map:
                        continue
                    new_val = edited_map[key]
                    target = "B" if ws[f'B{r}'].value in (None,"") else "C"
                    try:
                        ws[f"{target}{r}"].value = float(new_val)
                    except Exception:
                        ws[f"{target}{r}"].value = new_val
                    updates += 1
                st.success(f"✅ Excel updated: {updates} row(s) written.")

    # --- Download workbook ---
    to_save = io.BytesIO()
    wb.save(to_save)
    to_save.seek(0)
    st.download_button(
        label="💾 Download updated workbook",
        data=to_save,
        file_name="updated_" + uploaded_excel.name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
else:
    st.info("📂 Please upload your Excel template to begin.")
