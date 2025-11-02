import io
import re
import cv2
import numpy as np
import pytesseract
import pandas as pd
import openpyxl
import streamlit as st
from PIL import Image
from datetime import datetime
import os
import hashlib

# -----------------------------------------------------------
#  Streamlit Page Config
# -----------------------------------------------------------
st.set_page_config(page_title="Gas Analysis → Excel Populator", layout="wide")
st.title("Gas Analysis → Excel Populator (Local Template, Sticky Editor, B→C)")

st.markdown("""
**Flow**
1) App loads local template:  
   `/Users/fuxinghuang/Documents/Gasmet Reference Limits -Chemical Burning.xlsx`  
2) Click **🧹 Clear B & C** → creates a timestamped copy (keeps formats, keeps B1/C1).  
3) Upload image #1 → edit → **Populate** writes to **Column B**.  
4) Upload image #2 → edit → **Populate** writes to **Column C**.  
""")

# -----------------------------------------------------------
#  Constants & session init
# -----------------------------------------------------------
TEMPLATE_PATH = "/Users/fuxinghuang/Documents/Gasmet Reference Limits -Chemical Burning.xlsx"

if "working_path" not in st.session_state:
    st.session_state["working_path"] = TEMPLATE_PATH

if "populate_phase" not in st.session_state:
    st.session_state["populate_phase"] = "B"   # first populate -> B, second -> C

if "grid_df" not in st.session_state:
    st.session_state["grid_df"] = pd.DataFrame(columns=["Component", "Concentration"])

if "last_upload_id" not in st.session_state:
    st.session_state["last_upload_id"] = None

# -----------------------------------------------------------
#  Excel helpers
# -----------------------------------------------------------
def load_workbook(path):
    return openpyxl.load_workbook(path, data_only=False)

def clear_columns_B_C_keep_titles(ws):
    """Clear all non-formula values in columns B and C **except row 1** (keeps B1, C1)."""
    max_row = ws.max_row
    for r in range(2, max_row + 1):   # start at row 2 to keep titles
        for col in ("B", "C"):
            cell = ws[f"{col}{r}"]
            if isinstance(cell.value, str) and cell.value.strip().startswith("="):
                continue  # keep formulas
            cell.value = None

def save_copy_with_timestamp(base_path):
    """Save workbook at base_path as a new file with timestamp; return new_path."""
    wb = load_workbook(base_path)
    new_base, ext = os.path.splitext(base_path)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    new_path = f"{new_base}_{ts}{ext}"
    wb.save(new_path)
    return new_path

def write_map_to_column(ws, edited_map, target_column):
    """Write edited_map values into target_column (B or C) by matching Column A keys."""
    updates = 0
    for r in range(2, ws.max_row + 1):   # start at row 2; don't touch headers
        comp = ws[f"A{r}"].value
        if not comp:
            continue
        key = str(comp).strip().lower().replace(" ", "")
        if key not in edited_map:
            continue
        new_val = edited_map[key]
        try:
            ws[f"{target_column}{r}"].value = float(new_val)
        except Exception:
            ws[f"{target_column}{r}"].value = str(new_val)
        updates += 1
    return updates

# -----------------------------------------------------------
#  OCR (UNCHANGED as requested)
# -----------------------------------------------------------
def extract_table_from_image(image_bytes: bytes):
    """Detect table, deskew, OCR, and parse Component–Concentration pairs (auto invert + fallback)."""
    image_array = np.frombuffer(image_bytes, np.uint8)
    img = cv2.imdecode(image_array, cv2.IMREAD_COLOR)

    # ---------- Step 1: Preprocess ----------
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    gray = cv2.GaussianBlur(gray, (3, 3), 0)

    # Contrast enhancement
    gray = cv2.convertScaleAbs(gray, alpha=1.5, beta=10)

    # ---------- Step 2: Edge detection ----------
    edges = cv2.Canny(gray, 50, 150, apertureSize=3)

    # ---------- Step 3: Find largest rectangular contour ----------
    contours, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    table_contour = None
    max_area = 0
    for cnt in contours:
        area = cv2.contourArea(cnt)
        if area > max_area:
            peri = cv2.arcLength(cnt, True)
            approx = cv2.approxPolyDP(cnt, 0.02 * peri, True)
            if len(approx) == 4:
                table_contour = approx
                max_area = area

    # ---------- Step 4: Deskew (if contour found) ----------
    if table_contour is not None:
        pts = table_contour.reshape(4, 2)
        rect = np.zeros((4, 2), dtype="float32")
        s = pts.sum(axis=1)
        rect[0] = pts[np.argmin(s)]
        rect[2] = pts[np.argmax(s)]
        diff = np.diff(pts, axis=1)
        rect[1] = pts[np.argmin(diff)]
        rect[3] = pts[np.argmax(diff)]

        (tl, tr, br, bl) = rect
        widthA = np.linalg.norm(br - bl)
        widthB = np.linalg.norm(tr - tl)
        heightA = np.linalg.norm(tr - br)
        heightB = np.linalg.norm(tl - bl)
        maxWidth = int(max(widthA, widthB))
        maxHeight = int(max(heightA, heightB))

        dst = np.array([
            [0, 0],
            [maxWidth - 1, 0],
            [maxWidth - 1, maxHeight - 1],
            [0, maxHeight - 1]], dtype="float32")

        M = cv2.getPerspectiveTransform(rect, dst)
        warped = cv2.warpPerspective(img, M, (maxWidth, maxHeight))
    else:
        warped = img.copy()

    # ---------- Step 5: Threshold and check for inversion ----------
    gray = cv2.cvtColor(warped, cv2.COLOR_BGR2GRAY)
    gray = cv2.fastNlMeansDenoising(gray, h=25)

    # Adaptive threshold
    binary = cv2.adaptiveThreshold(
        gray, 255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY,
        31, 10
    )

    # Detect if background is dark (invert if needed)
    if np.mean(binary) < 127:
        binary = cv2.bitwise_not(binary)

    # Sharpen edges slightly
    kernel = np.array([[0, -1, 0],
                       [-1, 5, -1],
                       [0, -1, 0]])
    binary = cv2.filter2D(binary, -1, kernel)

    # ---------- Step 6: OCR ----------
    text = pytesseract.image_to_string(binary, config="--psm 6")

    # Fallback if empty text — try the original image
    if not text.strip():
        text = pytesseract.image_to_string(img, config="--psm 6")

    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]

    # ---------- Step 7: Parse results ----------
    table = {}
    for line in lines:
        # normalize separators
        line = line.replace("=", " ").replace(":", " ").replace(",", ".")
        m = re.search(r'([A-Za-z0-9\-\s/()]+?)\s+([0-9]+(?:\.[0-9]+)?)$', line)
        if m:
            comp = m.group(1).strip()
            conc = m.group(2).strip()
            table[comp] = conc
        else:
            nums = re.findall(r'[0-9]+(?:\.[0-9]+)?', line)
            if nums:
                conc = nums[-1]
                comp = line.replace(conc, "").strip()
                if len(comp) > 1:
                    table[comp] = conc

    return table, text

# -----------------------------------------------------------
#  Load template or working copy
# -----------------------------------------------------------
if not os.path.exists(TEMPLATE_PATH):
    st.error(f"❌ Template not found: {TEMPLATE_PATH}")
    st.stop()

working_path = st.session_state["working_path"]
if not os.path.exists(working_path):
    working_path = TEMPLATE_PATH
    st.session_state["working_path"] = working_path

wb = load_workbook(working_path)
ws = wb.active
st.success(f"📄 Active workbook: {working_path}  |  Sheet: {ws.title}")

# -----------------------------------------------------------
#  Clear B & C (keep titles B1/C1) -> create new copy & reset phase
# -----------------------------------------------------------
if st.button("🧹 Clear Columns B & C (keep B1/C1) & create timestamped copy"):
    # Always start from the original template for a clean copy
    base_wb = load_workbook(TEMPLATE_PATH)
    base_ws = base_wb.active
    clear_columns_B_C_keep_titles(base_ws)
    new_path = save_copy_with_timestamp(TEMPLATE_PATH)
    # save_copy_with_timestamp saved a clean copy of the original; now overwrite that file with our cleared base_wb
    base_wb.save(new_path)

    st.session_state["working_path"] = new_path
    st.session_state["populate_phase"] = "B"     # first populate goes to B
    st.session_state["grid_df"] = pd.DataFrame(columns=["Component", "Concentration"])
    st.session_state["last_upload_id"] = None
    st.success(f"🆕 New working file: {new_path} (B & C cleared; headers kept)")

# reload the (possibly new) working file
working_path = st.session_state["working_path"]
wb = load_workbook(working_path)
ws = wb.active

# -----------------------------------------------------------
#  Image upload -> OCR -> Editable grid (sticky)
# -----------------------------------------------------------
uploaded_img = st.file_uploader("📸 Upload Gasmet result image", type=["png", "jpg", "jpeg"])

if uploaded_img:
    image_bytes = uploaded_img.read()
    # build a stable id for this upload to detect changes
    upload_id = uploaded_img.name + "_" + hashlib.md5(image_bytes).hexdigest()

    # Run OCR only if new upload OR if grid is empty
    if upload_id != st.session_state["last_upload_id"]:
        st.info("Running local OCR…")
        table, raw_text = extract_table_from_image(image_bytes)
        st.text_area("📄 Raw OCR text (for review)", raw_text or "(no text detected)", height=240)

        # Initialize grid_df from OCR on first upload of this image
        df = pd.DataFrame(list(table.items()), columns=["Component", "Concentration"])
        if df.empty:
            df = pd.DataFrame(columns=["Component", "Concentration"])
        st.session_state["grid_df"] = df.copy()
        st.session_state["last_upload_id"] = upload_id
    else:
        # Same image reselected; don't reset user's edits
        st.info("Recognized same image upload — keeping your edits.")
        st.text_area("📄 Raw OCR text (for review)", "(unchanged)", height=100)

    st.markdown(f"### ✏️ Edit data to populate Column **{st.session_state['populate_phase']}**")
    edited_df = st.data_editor(
        st.session_state["grid_df"],
        num_rows="dynamic",
        use_container_width=True,
        key="grid_editor"
    )
    # Persist edits (this makes Enter/blur stick)
    st.session_state["grid_df"] = edited_df.copy()

    # -------------------------------------------------------
    #  Populate button → writes to B or C depending on phase
    # -------------------------------------------------------
    if st.button("🧾 Populate Excel with Edited Values"):
        ed = st.session_state["grid_df"]
        if ed is None or ed.empty:
            st.error("⚠️ No rows to populate — please add/edit rows above.")
        else:
            edited_map = {
                str(r["Component"]).strip().lower().replace(" ", ""): str(r["Concentration"])
                for _, r in ed.iterrows()
                if str(r.get("Component", "")).strip() != ""
            }
            target_col = st.session_state["populate_phase"]  # 'B' or 'C'
        
            # 🔥 FIX 1: Reload workbook to get current state
            try:
                from openpyxl import load_workbook
                wb = load_workbook(st.session_state["working_path"])
                ws = wb.active  # or wb[sheet_name] if you have a specific sheet
            
                # 🔥 FIX 2: Clear the target column first (assuming data starts at row 2)
                # Adjust the range based on your template structure
                max_row = ws.max_row
                for row_idx in range(2, max_row + 1):  # Start from row 2 (skip header)
                    cell = ws[f"{target_col}{row_idx}"]
                    cell.value = None  # Clear existing value
            
                # Now populate with new values
                updates = write_map_to_column(ws, edited_map, target_col)
            
                # Save the workbook
                wb.save(st.session_state["working_path"])
                wb.close()  # Important: close to release file lock
            
                st.success(f"✅ Populated Column {target_col}: {updates} row(s) saved → {st.session_state['working_path']}")
            
                # Switch to next phase
                if st.session_state["populate_phase"] == "B":
                    st.session_state["populate_phase"] = "C"
                    st.rerun()  # Force rerun to update UI
                
            except Exception as e:
                st.error(f"❌ Could not save workbook: {e}")
                st.info("Try closing the Excel file if it's open and rerun the step.")

# -----------------------------------------------------------
#  Download current working file
# -----------------------------------------------------------
if os.path.exists(st.session_state["working_path"]):
    with open(st.session_state["working_path"], "rb") as f:
        st.download_button(
            label="💾 Download current workbook",
            data=f.read(),
            file_name=os.path.basename(st.session_state["working_path"]),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
else:
    st.warning("No working file available yet. Click 'Clear B & C' to create one.")
